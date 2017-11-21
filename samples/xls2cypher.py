#!/usr/bin/env python3

from openpyxl import load_workbook

wb = load_workbook('sample1.xlsx', True)

sheetnames = wb.sheetnames

for wsname in sheetnames:
    ws=wb.get_sheet_by_name(wsname)
    #print ('Debug:', wsname, 'r:', ws.max_row, 'c:', ws.max_column)
    (graphelement,objectname) = ws.title.split('$')
    if graphelement == 'Node':
        for r in range(2, ws.max_row+1):
            print ('CREATE (', ws['A' + str(r)].value, ':',
                str(objectname), end='')
            if ws.max_column > 1:
                print (' { ', end='')
                for c in range(2, ws.max_column+1):
                    print (ws.cell(row=1, column=c).value, end='')
                    print (': \'', end='')
                    print (ws.cell(row=r, column=c).value, end='')
                    print ('\'', end='')
                    if c < ws.max_column:
                        print (', ', end='')
                print (' }', end='')
            print (')')
    elif graphelement == 'Relation':
        for r in range(2, ws.max_row+1):
            print ('CREATE (', ws.cell(column=1, row=r).value,
                ')-[:', objectname,']->(', ws.cell(column=2, row=r).value,
                ')', sep='')
